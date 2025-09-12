import pandas as pd
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
import os
import io
# Use relative imports for local modules if needed

def read_spielplan(file_bytes):
    """Read spielplan from Excel bytes"""
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name='Ergebnisse')
    return df

def create_spielbericht(match, template_bytes, template_placeholders, players_by_team=None):
    """Create a single match report with improved exception resilience, using template_placeholders for efficient replacement."""
    tf = None
    tmp_out = None
    result_bytes = None
    print("Creating match report...")
    print("Template placeholders:")
    for row, col, placeholder in template_placeholders:
        print(f" - {placeholder} (Cell: {row},{col})")
    try:
        tf = NamedTemporaryFile(delete=False, suffix='.xlsx')
        tf.write(template_bytes)
        tf.flush()
        try:
            wb = load_workbook(tf.name)
            ws = wb.active
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return None

        # Extract match data with error handling
        try:
            no = match.get('id', "")
            startzeit = match.get('Startzeit', "")
            datum = match.get('Tag', "")
            spielfeld = match.get('Feld', "")
            liga = match.get('Liga', "")
            gruppe = match.get('Gruppe', "")
            if pd.isna(gruppe):
                gruppe = ""
            team1 = match.get('Team 1', "")
            team2 = match.get('Team 2', "")
            schiedsrichter1 = match.get('Schiedsrichter', "")
            schiedsrichter2 = match.get('Schiedsrichter 2', "")
            if pd.isna(schiedsrichter2):
                schiedsrichter2 = ""
        except Exception as e:
            print(f"Error extracting match data: {e}")
            return None

        from openpyxl.styles import Alignment

        def safe_set_cell(row, col, value):
            try:
                cell = ws.cell(row, col)
                if hasattr(cell, 'value'):
                    cell.value = value
                    if isinstance(value, str) and '\n' in value:
                        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            except Exception as e:
                print(f"Error setting cell ({row},{col}): {e}")

        # Get player data for teams if available
        team1_players = []
        team2_players = []

        if players_by_team:
            # Try to find players for team1 and team2 in their respective Liga
            if liga in players_by_team:
                # Team 1 players
                if team1 in players_by_team[liga]:
                    team1_players = players_by_team[liga][team1]
                    print(f"Found {len(team1_players)} players for team '{team1}' in liga '{liga}'")
                else:
                    print(f"No players found for team '{team1}' in liga '{liga}'")

                # Team 2 players
                if team2 in players_by_team[liga]:
                    team2_players = players_by_team[liga][team2]
                    print(f"Found {len(team2_players)} players for team '{team2}' in liga '{liga}'")
                else:
                    print(f"No players found for team '{team2}' in liga '{liga}'")

        # Determine Spielzeit based on age category in liga or gruppe
        liga_gruppe = f"{liga} {gruppe}".lower()

        # Check for U12, u12, U-12 variations in both liga and gruppe
        if any(u12_pattern in liga_gruppe for u12_pattern in ["u12", "u-12", "u 12"]):
            duration = "2x7 Minuten"
            pause = "3 Minuten"

        # Check for U14, U16 variations
        elif any(pattern in liga_gruppe for pattern in ["u14", "u-14", "u 14", "u16", "u-16", "u 16"]):
            duration = "2x10 Minuten"
            pause = "5 Minuten"


        # Create the vermerk text with "GRUPPE"
        vermerk_text = f"{gruppe}"

        # Mapping from placeholder to value
        placeholder_values = {
            "$HEIM": team1,
            "$GEGNER": team2,
            "$SCHIRI": schiedsrichter1,
            "$SCHIRI2": schiedsrichter2,
            "$DATE": datum,
            "$TIME": startzeit,
            "$FIELD": spielfeld,
            "$LIGA": liga,
            "$NO": no,
            "$VERMERK": vermerk_text,
            "$DURATION": duration,
            "$PAUSE": pause
        }

        # Add player placeholders for Team 1 (Home)
        for i in range(1, 11):  # Player numbers 1-10
            player_idx = i - 1
            if player_idx < len(team1_players):
                placeholder_values[f"$NAMEH{i}"] = team1_players[player_idx].get("Name", "")
                nummer_h = team1_players[player_idx].get("Nummer", "")
                placeholder_values[f"$NH{i}"] = str(nummer_h) + " " if nummer_h else ""
            else:
                placeholder_values[f"$NAMEH{i}"] = ""
                placeholder_values[f"$NH{i}"] = ""

        # Add player placeholders for Team 2 (Away/Guest)
        for i in range(1, 11):  # Player numbers 1-10
            player_idx = i - 1
            if player_idx < len(team2_players):
                placeholder_values[f"$NAMEG{i}"] = team2_players[player_idx].get("Name", "")
                nummer_g = team2_players[player_idx].get("Nummer", "")
                placeholder_values[f"$NG{i}"] = str(nummer_g) + " " if nummer_g else ""
            else:
                placeholder_values[f"$NAMEG{i}"] = ""
                placeholder_values[f"$NG{i}"] = ""


        for row, col, placeholder in template_placeholders:
            value = placeholder_values.get(placeholder, None)
            try:
                if value is not None:
                    safe_set_cell(row, col, value)
                else:
                    safe_set_cell(row, col, "")
            except Exception as e:
                print(f"Error setting cell ({row},{col}): {e}")

        # Save to temporary file and return bytes
        try:
            tmp_out = NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(tmp_out.name)
            tmp_out.flush()
            with open(tmp_out.name, 'rb') as f:
                result_bytes = f.read()
        except Exception as e:
            print(f"Error saving or reading output file: {e}")
            return None

        return result_bytes
    except Exception as e:
        print(f"General error in create_spielbericht: {e}")
        return None
    finally:
        # Cleanup temp files
        if tf is not None:
            try:
                os.unlink(tf.name)
            except Exception as e:
                print(f"Error cleaning up temp file {tf.name}: {e}")
        if tmp_out is not None:
            try:
                os.unlink(tmp_out.name)
            except Exception as e:
                print(f"Error cleaning up temp file {tmp_out.name}: {e}")
