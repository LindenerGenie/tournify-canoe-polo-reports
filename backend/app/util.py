import re
from typing import List, Tuple
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
import numpy as np
import pandas as pd
import io

def find_placeholders_in_template(template_bytes) -> List[Tuple[int, int, str]]:
    """
    Scan the Excel template for placeholders matching $[a-z0-9]+ and return their positions and values.
    Returns a list of (row, col, placeholder) tuples.
    """
    tf = NamedTemporaryFile(delete=False, suffix='.xlsx')
    tf.write(template_bytes)
    tf.flush()
    try:
        wb = load_workbook(tf.name)
        ws = wb.active
        placeholder_pattern = re.compile(r"\$[A-Za-z0-9]+")
        placeholders = []
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row, col).value
                if isinstance(val, str) and placeholder_pattern.fullmatch(val):
                    placeholders.append((row, col, val))
        return placeholders
    finally:
        try:
            os.unlink(tf.name)
        except Exception:
            pass

def clean_json(obj):
    """Recursively replace NaN and infinite values with None for JSON serialization."""
    if isinstance(obj, dict):
        return {k: clean_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_json(v) for v in obj]
    elif isinstance(obj, float):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return obj
    return obj

def read_players_by_team(excel_bytes):
    """
    Reads an Excel file containing player assignments from multiple sheets and returns a dictionary:
    {
        "liga_name": {
            "team_name": [ { "Nummer": ..., "Name": ... }, ... ]
        }
    }
    Only teams with assigned players are included.
    Each sheet is expected to represent a different Liga.
    """
    xlsx = pd.ExcelFile(io.BytesIO(excel_bytes))
    players_by_liga_team = {}

    # Process each sheet in the Excel file
    for sheet_name in xlsx.sheet_names:
        df = pd.read_excel(xlsx, sheet_name=sheet_name)
        liga = sheet_name  # Use sheet name as Liga

        if liga not in players_by_liga_team:
            players_by_liga_team[liga] = {}

        # Expect columns: "Team", "Nummer", "Name"
        for _, row in df.iterrows():
            team = row.get("Team")
            nummer = row.get("Nummer")
            name = row.get("Name")

            if pd.isna(team) or pd.isna(nummer) or pd.isna(name):
                continue

            team = str(team)
            if team not in players_by_liga_team[liga]:
                players_by_liga_team[liga][team] = []

            players_by_liga_team[liga][team].append({
                "Nummer": int(nummer),
                "Name": str(name)
            })

    return players_by_liga_team
